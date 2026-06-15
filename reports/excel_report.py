from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette — dark-adjacent for screen, readable in Excel light mode
# ---------------------------------------------------------------------------
C_DARK_BG    = "1C2128"
C_MID_BG     = "21262D"
C_HEADER_BG  = "0D1117"
C_ACCENT     = "1F6FEB"
C_TEXT_LIGHT = "FFFFFF"
C_TEXT_MID   = "8B949E"
C_MALICIOUS  = "E74C3C"
C_SUSPICIOUS = "F39C12"
C_CLEAN      = "27AE60"
C_UNKNOWN    = "484F58"
C_BORDER     = "30363D"


_VERDICT_FILL = {
    "MALICIOUS":  PatternFill("solid", fgColor=C_MALICIOUS),
    "SUSPICIOUS": PatternFill("solid", fgColor=C_SUSPICIOUS),
    "CLEAN":      PatternFill("solid", fgColor=C_CLEAN),
    "UNKNOWN":    PatternFill("solid", fgColor=C_UNKNOWN),
}

_AUTH_FILL = {
    "pass":      PatternFill("solid", fgColor=C_CLEAN),
    "fail":      PatternFill("solid", fgColor=C_MALICIOUS),
    "softfail":  PatternFill("solid", fgColor=C_SUSPICIOUS),
    "neutral":   PatternFill("solid", fgColor=C_UNKNOWN),
    "permerror": PatternFill("solid", fgColor=C_MALICIOUS),
    "temperror": PatternFill("solid", fgColor=C_SUSPICIOUS),
    "none":      PatternFill("solid", fgColor=C_UNKNOWN),
}


def _h1(ws, row, col, text):
    """Section heading style."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=C_TEXT_LIGHT, size=11, name="Arial")
    cell.fill = PatternFill("solid", fgColor=C_MID_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell


def _col_header(ws, row, col, text):
    """Column header style."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=C_TEXT_LIGHT, size=9, name="Arial")
    cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _data_cell(ws, row, col, value, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(color="2D3748", size=9, name="Arial")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    return cell


def _kv_row(ws, row, label, value, label_color=C_TEXT_MID):
    l = ws.cell(row=row, column=1, value=label)
    l.font = Font(bold=True, color=label_color, size=9, name="Arial")
    l.alignment = Alignment(horizontal="left", vertical="top")

    v = ws.cell(row=row, column=2, value=str(value) if value else "N/A")
    v.font = Font(color="2D3748", size=9, name="Arial")
    v.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _overall_verdict(enriched):
    if not enriched:
        return "NO ENRICHMENT"
    malicious = sum(1 for r in enriched if r.get("verdict") == "MALICIOUS")
    suspicious = sum(1 for r in enriched if r.get("verdict") == "SUSPICIOUS")
    if malicious > 0:
        return "MALICIOUS"
    if suspicious > 0:
        return "SUSPICIOUS"
    return "CLEAN"


# ---------------------------------------------------------------------------
# Sheet 1 — Executive Summary
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, email_data, iocs, enriched, org, timestamp):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    headers = email_data.get("headers", {})
    auth = email_data.get("auth", {})

    overall = _overall_verdict(enriched)
    fill = _VERDICT_FILL.get(overall, PatternFill("solid", fgColor=C_UNKNOWN))

    # Title block
    title = ws.cell(row=1, column=1, value="PHISHING ANALYSIS REPORT")
    title.font = Font(bold=True, color=C_TEXT_LIGHT, size=16, name="Arial")
    title.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:B1")

    verdict_cell = ws.cell(row=2, column=1, value=f"VERDICT: {overall}")
    verdict_cell.font = Font(bold=True, color=C_TEXT_LIGHT, size=13, name="Arial")
    verdict_cell.fill = fill
    verdict_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 24

    ws.cell(row=3, column=1, value="")

    # Email metadata section
    _h1(ws, 4, 1, "Email Metadata")
    ws.merge_cells("A4:B4")
    ws.row_dimensions[4].height = 20

    meta = [
        ("From", headers.get("from", "")),
        ("From Address", headers.get("from_address", "")),
        ("Reply-To", headers.get("reply_to", "")),
        ("Return-Path", headers.get("return_path", "")),
        ("Subject", headers.get("subject", "")),
        ("Date", headers.get("date", "")),
        ("X-Originating-IP", headers.get("x_originating_ip", "") or "N/A"),
        ("X-Mailer", headers.get("x_mailer", "") or "N/A"),
    ]
    for i, (label, value) in enumerate(meta, start=5):
        _kv_row(ws, i, label, value)
        ws.row_dimensions[i].height = 16

    # Authentication section
    r = 5 + len(meta) + 1
    _h1(ws, r, 1, "Email Authentication")
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 20
    r += 1

    for proto in ("spf", "dkim", "dmarc"):
        result = auth.get(proto, "none")
        label_cell = ws.cell(row=r, column=1, value=proto.upper())
        label_cell.font = Font(bold=True, color=C_TEXT_MID, size=9, name="Arial")
        val_cell = ws.cell(row=r, column=2, value=result.upper())
        val_cell.font = Font(bold=True, color=C_TEXT_LIGHT, size=9, name="Arial")
        val_cell.fill = _AUTH_FILL.get(result, PatternFill("solid", fgColor=C_UNKNOWN))
        r += 1

    # IOC summary
    r += 1
    _h1(ws, r, 1, "IOC Summary")
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 20
    r += 1

    ioc_counts = [
        ("URLs Extracted", len(iocs.get("urls", []))),
        ("IPs Extracted", len(iocs.get("ips", []))),
        ("Domains Extracted", len(iocs.get("domains", []))),
        ("Hashes Extracted", len(iocs.get("hashes", []))),
        ("Total IOCs", sum(len(v) for v in iocs.values())),
    ]
    if enriched:
        malicious = sum(1 for e in enriched if e.get("verdict") == "MALICIOUS")
        suspicious = sum(1 for e in enriched if e.get("verdict") == "SUSPICIOUS")
        clean = sum(1 for e in enriched if e.get("verdict") == "CLEAN")
        ioc_counts += [
            ("IOCs — MALICIOUS", malicious),
            ("IOCs — SUSPICIOUS", suspicious),
            ("IOCs — CLEAN", clean),
        ]

    for label, value in ioc_counts:
        _kv_row(ws, r, label, value)
        r += 1

    # Footer
    r += 1
    foot = ws.cell(row=r, column=1, value=f"Generated: {timestamp}  |  shankar-bettadapura  |  {org}")
    foot.font = Font(color=C_TEXT_MID, size=8, italic=True, name="Arial")
    ws.merge_cells(f"A{r}:B{r}")


# ---------------------------------------------------------------------------
# Sheet 2 — Header Analysis
# ---------------------------------------------------------------------------

def _build_header_sheet(ws, email_data):
    ws.sheet_view.showGridLines = False

    col_widths = [22, 80]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _h1(ws, 1, 1, "Email Header Analysis")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22

    _col_header(ws, 2, 1, "Header Field")
    _col_header(ws, 2, 2, "Value")
    ws.row_dimensions[2].height = 18

    raw_headers = email_data.get("raw_headers", [])
    for i, (name, value) in enumerate(raw_headers, start=3):
        n = ws.cell(row=i, column=1, value=name)
        n.font = Font(bold=True, color="2D3748", size=9, name="Arial")
        n.alignment = Alignment(horizontal="left", vertical="top")

        v = ws.cell(row=i, column=2, value=str(value)[:500])
        v.font = Font(color="2D3748", size=9, name="Arial")
        v.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[i].height = 14

    # Received chain on next section
    r = len(raw_headers) + 4
    _h1(ws, r, 1, "Received Chain (Hop Analysis)")
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 22
    r += 1

    _col_header(ws, r, 1, "Hop #")
    _col_header(ws, r, 2, "Raw Header Value")
    r += 1

    for hop in email_data.get("received_chain", []):
        ws.cell(row=r, column=1, value=hop["hop"]).font = Font(bold=True, size=9, name="Arial")
        v = ws.cell(row=r, column=2, value=hop["raw"][:500])
        v.font = Font(size=9, name="Arial")
        v.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 14
        r += 1


# ---------------------------------------------------------------------------
# Sheet 3 — IOC Verdicts
# ---------------------------------------------------------------------------

def _build_ioc_sheet(ws, iocs, enriched):
    ws.sheet_view.showGridLines = False

    col_widths = [14, 60, 12, 14, 14, 40]
    headers_text = ["IOC Type", "Value (Defanged)", "Found In", "Verdict", "Sources Flagged", "TI Detail Summary"]
    for i, (w, h) in enumerate(zip(col_widths, headers_text), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _h1(ws, 1, 1, "IOC Enrichment Results")
    ws.merge_cells(f"A1:{get_column_letter(len(headers_text))}1")
    ws.row_dimensions[1].height = 22

    for col, text in enumerate(headers_text, start=1):
        _col_header(ws, 2, col, text)
    ws.row_dimensions[2].height = 18

    if not enriched:
        ws.cell(row=3, column=1, value="No enrichment data — run without --no-enrichment to query threat feeds.")
        return

    for i, r in enumerate(enriched, start=3):
        verdict = r.get("verdict", "UNKNOWN")
        fill = _VERDICT_FILL.get(verdict, PatternFill("solid", fgColor=C_UNKNOWN))

        _data_cell(ws, i, 1, r.get("ioc_type", ""))
        _data_cell(ws, i, 2, r.get("defanged", ""), wrap=True)
        _data_cell(ws, i, 3, r.get("ioc_source", ""))

        vd = ws.cell(row=i, column=4, value=verdict)
        vd.font = Font(bold=True, color=C_TEXT_LIGHT, size=9, name="Arial")
        vd.fill = fill
        vd.alignment = Alignment(horizontal="center", vertical="center")

        flagged = sum(1 for sr in r.get("source_results", []) if sr.get("malicious", False))
        total = len([sr for sr in r.get("source_results", []) if "error" not in sr])
        ws.cell(row=i, column=5, value=f"{flagged}/{total}").font = Font(size=9, name="Arial")

        # Summarise TI details
        parts = []
        for sr in r.get("source_results", []):
            if "error" in sr:
                parts.append(f'{sr["source"]}: ERROR')
            else:
                skip = {"source", "malicious", "error"}
                kv = ", ".join(f'{k.replace("_"," ").title()}: {v}' for k, v in sr.items() if k not in skip)
                parts.append(f'{sr["source"]}: {kv}')
        summary = " | ".join(parts)

        td = ws.cell(row=i, column=6, value=summary[:300])
        td.font = Font(size=9, name="Arial")
        td.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30


# ---------------------------------------------------------------------------
# Sheet 4 — Raw IOC Register
# ---------------------------------------------------------------------------

def _build_raw_ioc_sheet(ws, iocs):
    ws.sheet_view.showGridLines = False

    col_widths = [14, 60, 60, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _h1(ws, 1, 1, "Raw IOC Register")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 22

    for col, text in enumerate(["IOC Type", "Raw Value", "Defanged Value", "Found In"], start=1):
        _col_header(ws, 2, col, text)
    ws.row_dimensions[2].height = 18

    r = 3
    for ioc_type, key in [("URL", "urls"), ("IP", "ips"), ("Domain", "domains"), ("Hash", "hashes")]:
        for ioc in iocs.get(key, []):
            ws.cell(row=r, column=1, value=ioc_type).font = Font(bold=True, size=9, name="Arial")
            ws.cell(row=r, column=2, value=ioc.get("value", "")).font = Font(size=9, name="Arial", color="2D3748")
            ws.cell(row=r, column=3, value=ioc.get("defanged", "")).font = Font(size=9, name="Arial", color="2D3748")
            ws.cell(row=r, column=4, value=ioc.get("source", "")).font = Font(size=9, name="Arial", color="2D3748")
            for col in range(1, 5):
                ws.cell(row=r, column=col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = 14
            r += 1


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_excel_report(email_data: dict, iocs: dict, enriched: list, output_path: str, org: str = ""):
    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Executive Summary"
    _build_summary_sheet(ws1, email_data, iocs, enriched, org, timestamp)

    # Sheet 2
    ws2 = wb.create_sheet("Header Analysis")
    _build_header_sheet(ws2, email_data)

    # Sheet 3
    ws3 = wb.create_sheet("IOC Verdicts")
    _build_ioc_sheet(ws3, iocs, enriched)

    # Sheet 4
    ws4 = wb.create_sheet("Raw IOC Register")
    _build_raw_ioc_sheet(ws4, iocs)

    # Tab colours
    ws1.sheet_properties.tabColor = C_ACCENT
    ws2.sheet_properties.tabColor = "21262D"
    ws3.sheet_properties.tabColor = "21262D"
    ws4.sheet_properties.tabColor = "21262D"

    wb.save(output_path)